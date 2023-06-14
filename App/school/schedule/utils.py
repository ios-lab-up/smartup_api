from school import db
from school.tools.utils import *
from school.tools.utils import color
from school.models import Group, Schedule, Classroom, Teacher
from school.days.utils import *
from school.hours.utils import *
from datetime import datetime
from  typing import Optional
import traceback
import logging
import networkx as nx
import openpyxl
import os


def createSchedule(daysHours: list[str], classrooms: list[Classroom], group: Group) -> Schedule:
    '''Creates a schedule taking the days and hours from the schedule content and the classroom and groups objects'''
    try:

        if len(daysHours) != 0 and group:
            for i in range(len(daysHours)):
                schedule = Schedule(
                    day=daysHours[i].split()[0], # extract day from string using split
                    startTime=datetime.strptime(daysHours[i].split()[1], '%I:%M%p').strftime('%H:%M'), # extract start time from string and convert to 24-hour format
                    endTime=datetime.strptime(daysHours[i].split()[3], '%I:%M%p').strftime('%H:%M'), # extract end time from string and convert to 24-hour format
                    classroomID=classrooms[i]
                )

                db.session.add(schedule)
                db.session.commit()

                createScheduleGroupRelation(group, schedule)

        else:
            raise ValueError(
                f'{color(1,"Cannot create schedule: Does not contain timedate objs")} ❌'
            )

    except Exception as e:
        logging.critical(
            f'{color(5,"Schedule creation failed")} ❌: {e}\n{traceback.format_exc().splitlines()[-3]}')
        schedule = None

    return schedule



def createScheduleGroupRelation(group: Group, schedule: Schedule) -> None:
    '''Create a relation between schedule and group in DB'''
    try:
        if group and schedule:
            if schedule not in group.schedule:
                group.schedule.append(schedule)
                db.session.commit()
                logging.info(f'{color(4,"Schedule relation created")} ✅')
            else:
                logging.info(
                    f'{color(4,"Schedule relation already exists")} ✅')
        else:
            raise ValueError(
                f'{color(1,"Schedule relation creation failed")} ❌'
            )
    except Exception as e:
        logging.critical(
            f'{color(5,"Schedule relation creation failed")} ❌: {e}\n{traceback.format_exc().splitlines()[-3]}')



def getSchedule(Schedule: int) -> Schedule:
    '''Returns a object of type schedule given an id'''
    try:
        schedule = Schedule.query.filter_by(id=Schedule.id).first().toDict()
    except Exception as e:
        logging.critical(
            f'{color(5,"Schedule not found")} ❌: {e}\n{traceback.format_exc().splitlines()[-3]}')
        schedule = None

    return formatDateObjsSchedule(schedule)


def formatDateObjsSchedule(schedule: dict[str:str]) -> dict[str:str]:
    '''Formats the date objects in the schedule dictionary'''
    # Format the date objects in the dictionary
    schedule['creationDate'] = schedule['creationDate'].strftime(
        '%Y-%m-%d %H:%M:%S')
    schedule['lastupDate'] = schedule['lastupDate'].strftime(
        '%Y-%m-%d %H:%M:%S')
    schedule['startTime'] = schedule['startTime'].strftime("%H:%M:%S")
    schedule['endTime'] = schedule['endTime'].strftime("%H:%M:%S")
    return schedule

 
def createCompatibleSchedules(groups: list[dict[str,Group]], teachers: Optional[list[int]]=None, minimum: Optional[int]=3) -> list[list[dict[str,Group]]]:
    '''
    Returns a list of lists containing groups whose schedules don't overlap given a list of groups.
    If `teachers` is specified, only considers groups taught by those teachers.
    Only returns schedules that contain at least `minimum` groups.
    '''

    try:

        # Create a graph where each node represents a group
        graph = nx.Graph()
        for i in range(len(groups)):
            graph.add_node(i, group=groups[i])

        # Add an edge between two nodes if the corresponding groups have non-overlapping schedules and different subjects
        for i in range(len(groups)):
            for j in range(i+1, len(groups)):
                if not schedulesOverlap(groups[i], groups[j]) and groups[i]['subject'] != groups[j]['subject']:
                    graph.add_edge(i, j)

        # Find all cliques in the graph
        cliques = list(nx.find_cliques(graph))

        # Convert cliques to list of compatible schedules
        compatible_schedules = [list(map(lambda x: groups[x], clique)) for clique in cliques if len(clique) >= minimum]
        message = f"{len(compatible_schedules)} compatible schedules were found"      
        error = None
        status_code = 201


        # Filter schedules by teacher if the fitler is not none
        if teachers:
            teacher_names = []
            for teacher_id in teachers:
                teacher = Teacher.query.filter_by(id=teacher_id).first()
                if not teacher:
                    raise ValueError(color(3,f"Teacher with id: {teacher_id} does not exist"))
                else:
                    teacher_names.append(teacher.name)

            # Keep only those schedules that have at least one group taught by one of the teachers
            compatible_schedules = [schedule for schedule in compatible_schedules if any(group['teacher'] in teacher_names for group in schedule)]
            message = f"{len(compatible_schedules)} compatible schedules were found (filtered by teachers: {', '.join(teacher_names)})"


    except Exception as e:
        logging.error(color(1,f"Error creating compatible schedules: {e}"))
        
        compatible_schedules = []
        message = "Error creating compatible schedules"
        status_code = 500
        error = str(e)
    cleanSchedulesOutput(compatible_schedules)
    return compatible_schedules, message, status_code, error



def schedulesOverlap(group_1: Group, group_2: Group) -> bool:
    '''
    Returns True if the schedules of group_1 and group_2 overlap, False otherwise
    '''

    try:
        # Get the schedules of each group
        group_1_schedules = group_1['schedules']
        group_2_schedules = group_2['schedules']

        # Iterate over each schedule in group 1
        for schedule_1 in group_1_schedules:
            # Iterate over each schedule in group 2
            for schedule_2 in group_2_schedules:
                # Check if the days are the same
                if schedule_1['day'] == schedule_2['day']:
                    # Check if the times overlap
                    if (
                        schedule_1['startTime'] <= schedule_2['endTime'] and
                        schedule_1['endTime'] >= schedule_2['startTime']
                    ):
                        return True

        # If no overlap was found, return False
        return False

    except Exception as e:
        logging.critical(f'Schedule comparison failed: {e}')
        return False
    

def cleanSchedulesOutput(schedules):
    '''
    Cleans the schedules output in the next format
    List of dictionaries, each dictionary is a schedule
    Each schedule has a key for each day of the week
    Each day of the week has a list of dictionaries, each dictionary is a block of time
    '''
    cleanSchedules = []
    for schedule in schedules:
        Horario = {
            'Lunes': [],
            'Martes': [],
            'Miercoles': [],
            'Jueves': [],
            'Viernes': []
        }
        colorblockoptions = {
        "Morado": "c778ff",
        "Amarillo": "fffd78",
        "Rojo": "ff8178",
        "Verde": "95ff78",
        "Azul": "78fffa",
        "Naranja": "ffc478",
        "Rosa": "ff78db",
        "Verde Claro": "78ffae",
        "Azul Obscuro": "787fff",
        "Gris":"bab8ba"
        }
        for group in schedule:
            #select a random color for the group without repeating
            color=colorblockoptions.popitem()[1]
            for block in group['schedules']:
                if len(block['startTime']) == 8:
                    block['startTime'] = block['startTime'][:-3]
                    block['endTime'] = block['endTime'][:-3]
                hora=block['startTime']+' - '+block['endTime']
                if block['day']=='Lun':
                    Horario['Lunes'].append({hora:{'Clase': group['classNumber'],'Materia': group['subject'],'Maestro': group['teacher'], 'Salon': block['classroomID'], 'Color': color}})
                elif block['day']=='Mart':
                    Horario['Martes'].append({hora:{'Clase': group['classNumber'],'Materia': group['subject'],'Maestro': group['teacher'], 'Salon': block['classroomID'], 'Color': color}})
                elif block['day']=='Miérc':
                    Horario['Miercoles'].append({hora:{'Clase': group['classNumber'],'Materia': group['subject'],'Maestro': group['teacher'], 'Salon': block['classroomID'], 'Color': color}})
                elif block['day']=='Jue':
                    Horario['Jueves'].append({hora:{'Clase': group['classNumber'],'Materia': group['subject'],'Maestro': group['teacher'], 'Salon': block['classroomID'], 'Color': color}})
                elif block['day']=='V':
                    Horario['Viernes'].append({hora:{'Clase': group['classNumber'],'Materia': group['subject'],'Maestro': group['teacher'], 'Salon': block['classroomID'], 'Color': color}})
        cleanSchedules.append(Horario)
    if len(cleanSchedules) > 0:
        excelOutput(cleanSchedules)
    return cleanSchedules

def excelOutput(schedules):
    print("Creating excel with schedules")
    wb = openpyxl.Workbook() # Create a blank workbook
    hourrowrelation={
        '07:00': 2,
        '07:30': 3,
        '08:00': 4,
        '08:30': 5,
        '09:00': 6,
        '09:30': 7,
        '10:00': 8,
        '10:30': 9,
        '11:00': 10,
        '11:30': 11,
        '12:00': 12,
        '12:30': 13,
        '13:00': 14,
        '13:30': 15,
        '14:00': 16,
        '14:30': 17,
        '15:00': 18,
        '15:30': 19,
        '16:00': 20,
        '16:30': 21,
        '17:00': 22,
        '17:30': 23,
        '18:00': 24,
        '18:30': 25,
        '19:00': 26,
        '19:30': 27,
        '20:00': 28,
        '20:30': 29,
        '21:00': 30,
        '21:30': 31,
    }
    schedulecount = 1
    for schedule in schedules:
        sheet=wb.create_sheet(title="Horario " + str(schedulecount))
        sheet = wb['Horario ' + str(schedulecount)]
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 35
        sheet.column_dimensions['C'].width = 35
        sheet.column_dimensions['D'].width = 35
        sheet.column_dimensions['E'].width = 35
        sheet.column_dimensions['F'].width = 35
        #change the height of the rows 2:32 to 30 pixels
        for i in range(2, 32):
            sheet.row_dimensions[i].height = 30
        sheet['A1'] = "Hora"
        sheet['B1'] = 'Lunes'
        sheet['C1'] = 'Martes'
        sheet['D1'] = 'Miércoles'
        sheet['E1'] = 'Jueves'
        sheet['F1'] = 'Viernes'
        # Fill from A2 to A31 with blocks of 30 minutes starting at 7:00 (7:00 - 7:30, 7:30 - 8:00, etc.)
        for i in range(2, 32):
            sheet['A' + str(i)] = f"{6 + i // 2}:{30 * (i % 2):02d} - {6 + (i + 1) // 2}:{30 * ((i + 1) % 2):02d}"
        days = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']
        row = 2
        for day in days:
            if day in schedule:
                classes = schedule[day]
                for class_info in classes:
                    for class_period, details in class_info.items():
                        #align the block of time with the corresponding row using the hourrowrelation dictionary
                        start_time, end_time = class_period.split(' - ')
                        start_cell = sheet.cell(row=hourrowrelation[start_time], column=days.index(day) + 2) #
                        end_cell = sheet.cell(row=hourrowrelation[end_time]-1, column=days.index(day) + 2)
                        merge_range = start_cell.coordinate + ':' + end_cell.coordinate
                        sheet.merge_cells(merge_range)
                        #fill the block of time with the color of the class
                        sheet[start_cell.coordinate].fill = openpyxl.styles.PatternFill(start_color=details['Color'], end_color=details['Color'], fill_type='solid')
                        sheet[start_cell.coordinate] = f"{start_time} - {end_time} {details['Materia']} {details['Maestro']} ID Salon: {details['Salon']}"
                        sheet[start_cell.coordinate].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')                       
        #make the cells A1:F31 have a border
        for i in range(1, 32):
            for j in range(1, 7):
                sheet.cell(row=i, column=j).border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))
        #make the cells B2:F31 adjust the text to fit in the cell and center it horizontally and vertically
        for i in range(2, 32):
            for j in range(1, 7):
                sheet.cell(row=i, column=j).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        #fill the cells A1:F1 with a yellow background

        for i in range(1, 7):
            sheet.cell(row=1, column=i).fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        schedulecount += 1
    #delete default sheet
    wb.remove(wb['Sheet'])
    #save the workbook in static folder for docker container
    #check if the file exists and delete it
    if os.path.exists('App/school/static/schedules/PosibleSchedules.xlsx'):
        os.remove('App/school/static/schedules/PosibleSchedules.xlsx')
    path = 'App/school/static/schedules'
    wb.save(path + '/PosibleSchedules.xlsx')
